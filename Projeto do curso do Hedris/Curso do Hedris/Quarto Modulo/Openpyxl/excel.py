"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""

'''pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos ['paginas']

for linha in range(5, 100):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).valeue = 1200 + linha

    preco = round(uniform(10,200), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')'''

import openpyxl
from random import uniform

planilha = openpyxl.workbook()


planilha.creat_sheet('Planilha1', 0)
planilha.creat_sheet('Planilha2', 0)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1

    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha


    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco


for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha2.cell(linha, 1).value = numero_pedido
    planilha2.cell(linha, 2).value = 1200 + linha


    preco = round(uniform(10, 100), 2)
    planilha2.cell(linha, 3).value = preco